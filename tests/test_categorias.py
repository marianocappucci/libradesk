"""
Categorías de incidencia — catálogo de dos niveles, global (2026-08-02,
pendiente 20).

El pedido era literal: *"Tipo: Hardware → Impresoras"*. Hasta ahora eso vivía
dentro del título del ticket, así que "cuántas fallas de impresoras hubo este
mes" no se podía contestar sin leer los títulos uno por uno.

Lo que fijan estos tests, en orden de lo que se rompe sin que se note:

1. Que la ruta exista y no sea el fallback de la SPA.
2. Que el catálogo **no** admita un tercer nivel — es la invariante de la que
   dependen la UI y el reporte.
3. Que borrar una categoría **no borre tickets**, ni deje `categoria_id`
   apuntando a un id inexistente (el pragma de FKs está apagado).
4. Que filtrar por una categoría **raíz** traiga las de sus subcategorías.
"""
import pytest

RUTA = "/api/categorias"

# `client` sale de conftest.py.


def _login(client) -> None:
    assert client.post("/auth/login", json={"username": "admin", "password": "admin"}).status_code == 200


def _crear(client, nombre, parent_id=None):
    r = client.post(RUTA, json={"nombre": nombre, "parent_id": parent_id})
    assert r.status_code == 201, r.text
    return r.json()


@pytest.fixture
def catalogo(client):
    """Hardware → {Impresoras, Notebooks} · Software → {Sistema operativo}."""
    _login(client)
    hardware = _crear(client, "Hardware")
    software = _crear(client, "Software")
    return {
        "hardware": hardware["id"],
        "software": software["id"],
        "impresoras": _crear(client, "Impresoras", hardware["id"])["id"],
        "notebooks": _crear(client, "Notebooks", hardware["id"])["id"],
        "so": _crear(client, "Sistema operativo", software["id"])["id"],
    }


# ── Que la ruta exista ──────────────────────────────────────────────────────

def test_la_ruta_existe_de_verdad(client):
    """`asgi.py` sirve la SPA en `/{full_path:path}`: una ruta inventada
    devuelve 200 con HTML. Sin este chequeo el archivo entero podría estar
    midiendo el fallback."""
    assert RUTA in client.app.openapi()["paths"]


def test_sin_sesion_no_se_ve(client):
    assert client.get(RUTA).status_code == 401


# ── La forma del catálogo ───────────────────────────────────────────────────

def test_el_listado_viene_ordenado_como_arbol(client, catalogo):
    filas = client.get(RUTA).json()

    # Cada raíz seguida de sus hijas, en ese orden — es lo que consumen tanto
    # el `<select>` como la pantalla del ABM.
    assert [f["ruta"] for f in filas] == [
        "Hardware",
        "Hardware · Impresoras",
        "Hardware · Notebooks",
        "Software",
        "Software · Sistema operativo",
    ]
    impresoras = next(f for f in filas if f["nombre"] == "Impresoras")
    assert impresoras["parent_id"] == catalogo["hardware"]
    assert impresoras["parent_nombre"] == "Hardware"
    assert next(f for f in filas if f["nombre"] == "Hardware")["parent_nombre"] is None


def test_no_se_puede_anidar_un_tercer_nivel(client, catalogo):
    """La invariante de la que cuelga todo lo demás: la UI dibuja dos niveles y
    el reporte agrupa con **un** join contra sí misma."""
    r = client.post(RUTA, json={"nombre": "Láser", "parent_id": catalogo["impresoras"]})
    assert r.status_code == 400
    assert "dos niveles" in r.json()["detail"]


def test_padre_inexistente_es_404(client):
    _login(client)
    assert client.post(RUTA, json={"nombre": "X", "parent_id": 9999}).status_code == 404


def test_el_nombre_es_unico_dentro_del_mismo_padre(client, catalogo):
    assert client.post(RUTA, json={"nombre": "Impresoras", "parent_id": catalogo["hardware"]}).status_code == 409
    # Pero "Otros" puede existir bajo Hardware y bajo Software a la vez, que es
    # lo natural en un catálogo de dos niveles.
    assert client.post(RUTA, json={"nombre": "Otros", "parent_id": catalogo["hardware"]}).status_code == 201
    assert client.post(RUTA, json={"nombre": "Otros", "parent_id": catalogo["software"]}).status_code == 201


def test_renombrar_no_cambia_de_padre(client, catalogo):
    r = client.put(f"{RUTA}/{catalogo['impresoras']}", json={"nombre": "Impresoras y scanners"})
    assert r.status_code == 200
    assert r.json()["ruta"] == "Hardware · Impresoras y scanners"
    assert r.json()["parent_id"] == catalogo["hardware"]


# ── Clasificar un ticket ────────────────────────────────────────────────────

@pytest.fixture
def tickets(client, catalogo):
    cliente_id = client.post("/api/clientes", json={"nombre": "C", "email": "c@t.com"}).json()["id"]

    def ticket(titulo, categoria_id):
        return client.post("/api/incidencias", json={
            "cliente_id": cliente_id, "titulo": titulo, "categoria_id": categoria_id,
        }).json()["id"]

    return {
        "cliente_id": cliente_id,
        "impresora": ticket("No imprime", catalogo["impresoras"]),
        "notebook": ticket("No enciende", catalogo["notebooks"]),
        "so": ticket("Windows no actualiza", catalogo["so"]),
        "sin_clasificar": ticket("Consulta general", None),
    }


def test_la_incidencia_guarda_su_categoria(client, catalogo, tickets):
    inc = client.get(f"/api/incidencias/{tickets['impresora']}").json()
    assert inc["categoria_id"] == catalogo["impresoras"]
    assert client.get(f"/api/incidencias/{tickets['sin_clasificar']}").json()["categoria_id"] is None


def test_filtrar_por_categoria_hoja(client, catalogo, tickets):
    filtradas = client.get(f"/api/incidencias?categoria_id={catalogo['impresoras']}").json()
    assert [i["id"] for i in filtradas] == [tickets["impresora"]]


def test_el_filtro_de_la_api_es_exacto_no_recursivo(client, catalogo, tickets):
    """Documenta la diferencia con el reporte a propósito: `/api/incidencias`
    filtra por la categoría **exacta**; quien quiere el árbol entero es el
    reporte (y la lista del frontend, que lo resuelve con el catálogo que ya
    tiene cargado)."""
    assert client.get(f"/api/incidencias?categoria_id={catalogo['hardware']}").json() == []


# ── Borrado ─────────────────────────────────────────────────────────────────

def test_no_se_borra_una_categoria_con_subcategorias(client, catalogo):
    r = client.delete(f"{RUTA}/{catalogo['hardware']}")
    assert r.status_code == 409
    assert "2 subcategorías" in r.json()["detail"]


def test_no_se_borra_una_categoria_en_uso_sin_forzar(client, catalogo, tickets):
    r = client.delete(f"{RUTA}/{catalogo['impresoras']}")
    assert r.status_code == 409
    assert "1 incidencias" in r.json()["detail"]
    # Y no se borró nada: ni la categoría ni el ticket.
    assert any(c["id"] == catalogo["impresoras"] for c in client.get(RUTA).json())
    assert client.get(f"/api/incidencias/{tickets['impresora']}").json()["categoria_id"] == catalogo["impresoras"]


def test_forzar_desclasifica_los_tickets_pero_no_los_borra(client, catalogo, tickets):
    """El `ondelete` del modelo no corre —el pragma está apagado—, así que sin
    el `UPDATE` explícito el ticket quedaría apuntando a un id inexistente.
    Es el mismo hallazgo que este producto ya pagó con `equipos_movimientos`."""
    assert client.delete(f"{RUTA}/{catalogo['impresoras']}?forzar=true").status_code == 204

    inc = client.get(f"/api/incidencias/{tickets['impresora']}").json()
    assert inc["id"] == tickets["impresora"]          # el ticket sigue vivo
    assert inc["categoria_id"] is None                # y no colgado de un id muerto
    assert not any(c["id"] == catalogo["impresoras"] for c in client.get(RUTA).json())
    # Los otros tickets no se tocaron.
    assert client.get(f"/api/incidencias/{tickets['notebook']}").json()["categoria_id"] == catalogo["notebooks"]


def test_borrar_una_categoria_vacia_no_necesita_forzar(client, catalogo):
    assert client.delete(f"{RUTA}/{catalogo['so']}").status_code == 204


def test_categoria_inexistente_es_404(client):
    _login(client)
    assert client.delete(f"{RUTA}/9999").status_code == 404


# ── El reporte ──────────────────────────────────────────────────────────────

def test_el_reporte_acepta_el_filtro_por_categoria(client, catalogo, tickets):
    """Acá el filtro **sí** es recursivo: elegir "Hardware" tiene que contestar
    por impresoras y notebooks juntas, que es la pregunta que se hace de
    verdad."""
    from io import BytesIO

    from openpyxl import load_workbook

    def titulos(qs: str) -> set[str]:
        r = client.get(f"/api/reportes/incidencias-periodo.xlsx?{qs}")
        assert r.status_code == 200
        assert r.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml"
        )
        ws = load_workbook(BytesIO(r.content)).active
        # La columna 5 es "Título" (1 #, 2 Cliente, 3 Sector, 4 Categoría).
        return {
            fila[4] for fila in ws.iter_rows(values_only=True)
            if fila[0] and isinstance(fila[0], int)
        }

    periodo = "desde=2020-01-01&hasta=2099-12-31"
    assert titulos(f"{periodo}&categoria_id={catalogo['impresoras']}") == {"No imprime"}
    assert titulos(f"{periodo}&categoria_id={catalogo['hardware']}") == {"No imprime", "No enciende"}
    # Sin filtro entran las cuatro, incluida la que no tiene categoría.
    assert len(titulos(periodo)) == 4


# ── La migración ────────────────────────────────────────────────────────────

# El schema anterior de `incidencias` estaba escrito acá como una constante,
# para reconstruir la tabla a mano antes de correr la migración. Se fue el
# 2026-08-03 con el paso a Alembic: `downgrade` deja ese mismo estado sin que
# haya que transcribirlo, y una transcripción es justo lo que se desactualiza en
# silencio cuando el modelo cambia. El schema real de producción sí sigue
# escrito, como ancla de regresión, en tests/test_alembic.py.


def test_la_migracion_agrega_la_categoria_a_una_base_vieja(client, catalogo, tickets):
    """El caso real del deploy: `compulibra` tiene sus 23 incidencias desde la
    migración del Node.js, y el schema propio ya no lo crea `create_all()` sino
    la cadena de Alembic — sin la revisión 0002, `categoria_id` jamás llegaría
    ahí.

    La base se lleva al baseline con un `downgrade` real, que es lo que deja el
    estado exacto anterior a la revisión —incluida la desaparición de la tabla
    `categorias_incidencia`— y de paso ejercita el camino inverso. Reconstruir
    la tabla a mano no alcanzaba: `categoria_id` participa de una FK, así que
    `DROP COLUMN` no sirve, y retroceder sólo el número de versión dejaba la
    tabla del catálogo en pie y hacía fallar el `create_table` del upgrade.

    Esto verifica el efecto por la API; el mecanismo de la adopción está
    cubierto en tests/test_alembic.py.
    """
    from alembic import command
    from sqlalchemy import text

    from app import database
    from app.schema import BASELINE, _config, ensure_schema

    engine = database.get_engine()

    # Introspección por el inspector de SQLAlchemy y no por `PRAGMA
    # table_info` / `index_list`: eso es exclusivo de SQLite y la suite corre
    # también contra PostgreSQL (ver `tests/conftest.py`).
    from sqlalchemy import inspect

    def columnas_de(conn):
        return {c["name"] for c in inspect(conn).get_columns("incidencias")}

    with engine.begin() as conn:
        filas_antes = conn.execute(text("SELECT COUNT(*) FROM incidencias")).scalar()
        command.downgrade(_config(conn), BASELINE)
        columnas = columnas_de(conn)
    assert "categoria_id" not in columnas
    assert filas_antes == 4  # los 4 tickets del fixture

    assert ensure_schema(engine) == "upgrade"

    with engine.begin() as conn:
        columnas = columnas_de(conn)
        indices = {i["name"] for i in inspect(conn).get_indexes("incidencias")}
        filas_despues = conn.execute(text("SELECT COUNT(*) FROM incidencias")).scalar()
    assert "categoria_id" in columnas
    # Mismo nombre que genera SQLAlchemy: una base migrada y una nueva tienen
    # que quedar idénticas.
    assert "ix_incidencias_categoria_id" in indices
    # Los tickets migrados no se pierden ni se tocan: quedan sin clasificar.
    assert filas_despues == filas_antes
    assert client.get(f"/api/incidencias/{tickets['impresora']}").json()["categoria_id"] is None

    assert ensure_schema(engine) == "upgrade"  # idempotente

    # Y la base migrada acepta escribir la columna nueva, que es el punto.
    #
    # 🔴 La categoría se crea DE NUEVO acá, después del upgrade. Las de
    # `catalogo` ya no existen: el `downgrade` al baseline borra la tabla
    # `categorias_incidencia` entera. Hasta el 2026-08-09 este bloque reusaba
    # `catalogo["notebooks"]` y pasaba igual **porque SQLite no exige las
    # foreign keys** — el ticket quedaba apuntando a una categoría borrada y el
    # test lo daba por bueno. Contra PostgreSQL sale
    # `ForeignKeyViolation: Key (categoria_id)=(4) is not present`.
    notebooks = client.post("/api/categorias", json={"nombre": "Notebooks"}).json()["id"]
    inc = client.get(f"/api/incidencias/{tickets['impresora']}").json()
    r = client.put(f"/api/incidencias/{tickets['impresora']}", json={
        **{k: inc[k] for k in ("cliente_id", "titulo", "estado", "prioridad")},
        "categoria_id": notebooks,
    })
    assert r.json()["categoria_id"] == notebooks
    assert client.get(f"/api/incidencias?categoria_id={notebooks}").json()


def test_el_reporte_trae_la_ruta_completa_de_la_categoria(client, catalogo, tickets):
    from app import database
    from app.services.reportes import ReportesService

    filas = ReportesService(database.get_session_factory()).incidencias(
        desde="2020-01-01", hasta="2099-12-31",
    )
    por_titulo = {f["titulo"]: f["categoria"] for f in filas}
    assert por_titulo["No imprime"] == "Hardware · Impresoras"
    assert por_titulo["Consulta general"] is None
