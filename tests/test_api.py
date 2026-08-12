"""Smoke tests de la API real (auth + un flujo CRUD por dominio +
dashboard). No es cobertura exhaustiva de los 5 dominios — verifica que
el wiring completo (libraauth + SQLAlchemy + routers) funciona de punta
a punta contra una SQLite temporal."""
import pytest

# La fixture `client` vive en `conftest.py`: es la misma para toda la suite y
# ya no reimporta `app.*` en cada test. Ver ahí el porqué.


def _login(client) -> None:
    r = client.post("/auth/login", json={"username": "admin", "password": "admin"})
    assert r.status_code == 200


def test_health(client):
    assert client.get("/api/health").status_code == 200


def test_login_and_me(client):
    _login(client)
    r = client.get("/auth/me")
    assert r.status_code == 200
    assert r.json()["role"] == "admin"


def test_cliente_equipo_incidencia_flow(client):
    _login(client)

    r = client.post("/api/clientes", json={"nombre": "Cliente Test", "email": "t@test.com"})
    assert r.status_code == 201
    cliente_id = r.json()["id"]

    r = client.post("/api/equipos", json={"cliente_id": cliente_id, "tipo": "Notebook"})
    assert r.status_code == 201
    equipo_id = r.json()["id"]

    r = client.post("/api/tecnicos", json={"nombre": "Tecnico Test"})
    assert r.status_code == 201
    tecnico_id = r.json()["id"]

    r = client.post("/api/incidencias", json={
        "cliente_id": cliente_id, "equipo_id": equipo_id, "tecnico_id": tecnico_id,
        "titulo": "No enciende", "prioridad": "alta",
    })
    assert r.status_code == 201
    incidencia_id = r.json()["id"]
    assert r.json()["estado"] == "abierto"

    r = client.put(f"/api/incidencias/{incidencia_id}", json={
        "cliente_id": cliente_id, "equipo_id": equipo_id, "tecnico_id": tecnico_id,
        "titulo": "No enciende", "prioridad": "alta", "estado": "resuelta",
        "resolucion": "Fuente cambiada",
    })
    assert r.status_code == 200
    assert r.json()["fecha_cierre"] is not None

    log = client.get(f"/api/incidencias/{incidencia_id}/estados").json()
    assert len(log) == 2
    assert log[0]["estado_nuevo"] == "resuelta"

    dash = client.get("/api/dashboard").json()
    assert dash["incidencias_por_estado"].get("resuelta") == 1
    assert dash["total_clientes_activos"] == 1


def test_dashboard_prioridad_abiertas(client):
    """`incidencias_por_prioridad_abiertas` no tenia ningun test: se podia
    romper el filtro de estados y toda la suite seguia en verde. Aparecio al
    verificar la ficha del cliente (2026-08-02), que comparte con el resumen
    global la constante `ESTADOS_ABIERTOS` — si se rompe ahi, se rompen los
    dos y este es el unico que lo ve del lado global."""
    _login(client)
    cliente_id = client.post("/api/clientes", json={"nombre": "C", "email": "c@t.com"}).json()["id"]

    for titulo, prioridad in (("Urgente", "alta"), ("Otra", "alta"), ("Menor", "baja")):
        client.post("/api/incidencias", json={
            "cliente_id": cliente_id, "titulo": titulo, "prioridad": prioridad,
        })
    # Una cerrada, que NO tiene que contarse aunque sea de prioridad alta.
    cerrada = client.post("/api/incidencias", json={
        "cliente_id": cliente_id, "titulo": "Cerrada", "prioridad": "alta",
    }).json()["id"]
    client.put(f"/api/incidencias/{cerrada}", json={
        "cliente_id": cliente_id, "titulo": "Cerrada", "prioridad": "alta", "estado": "cerrado",
    })

    dash = client.get("/api/dashboard").json()
    assert dash["incidencias_por_prioridad_abiertas"] == {"alta": 2, "baja": 1}


def test_cliente_guarda_cuit_y_domicilio(client):
    """Pendiente 16 (2026-08-02): antes `clientes` solo tenia `ciudad`, asi
    que los dos datos fiscales se tipeaban **en cada comprobante** aunque
    fueran siempre los mismos."""
    _login(client)
    r = client.post("/api/clientes", json={
        "nombre": "Fiscal", "email": "f@t.com",
        "cuit": "30-71234567-8", "domicilio": "Av. Siempreviva 742",
    })
    assert r.status_code == 201
    cliente_id = r.json()["id"]
    assert r.json()["cuit"] == "30-71234567-8"
    assert r.json()["domicilio"] == "Av. Siempreviva 742"

    # Sobreviven al GET y se pueden vaciar.
    assert client.get(f"/api/clientes/{cliente_id}").json()["cuit"] == "30-71234567-8"
    r = client.put(f"/api/clientes/{cliente_id}", json={
        "nombre": "Fiscal", "email": "f@t.com", "cuit": None, "domicilio": None,
    })
    assert r.json()["cuit"] is None

    # Y son opcionales: un cliente sin datos fiscales se crea igual.
    r = client.post("/api/clientes", json={"nombre": "Sin datos", "email": "s@t.com"})
    assert r.status_code == 201
    assert r.json()["cuit"] is None and r.json()["domicilio"] is None


def test_no_entran_dos_clientes_con_el_mismo_cuit(client):
    """Lo que LibraDesk gana al compartir la tabla `clients` del motor: la
    validacion es la misma funcion que usa `create_client()` de LibraCore
    (`validar_cuit_no_duplicado`), no una copia.

    El CUIT se compara **sin guiones**, asi que el segundo alta choca aunque
    venga tipeado distinto -- que es el caso real: nadie lo escribe igual dos
    veces.
    """
    _login(client)
    assert client.post("/api/clientes", json={
        "nombre": "Compulibra", "email": "uno@t.com", "cuit": "30-71234567-8",
    }).status_code == 201

    r = client.post("/api/clientes", json={
        "nombre": "Compulibra otra vez", "email": "dos@t.com", "cuit": "30712345678",
    })
    assert r.status_code == 409
    # El mensaje dice de quien es y en que estado esta: un 409 pelado deja al
    # usuario sabiendo que no puede, y no por que.
    assert "Compulibra" in r.json()["detail"]

    # Un CUIT libre entra sin problema.
    assert client.post("/api/clientes", json={
        "nombre": "Otro", "email": "tres@t.com", "cuit": "27-99999999-4",
    }).status_code == 201


def test_editar_un_cliente_no_choca_con_su_propio_cuit(client):
    """Sin `excluir_id` en la validacion, guardarle el nombre a un cliente que
    tiene CUIT fallaria siempre contra su propia fila."""
    _login(client)
    cid = client.post("/api/clientes", json={
        "nombre": "Compulibra", "email": "c@t.com", "cuit": "30-71234567-8",
    }).json()["id"]

    r = client.put(f"/api/clientes/{cid}", json={
        "nombre": "Compulibra SRL", "email": "c@t.com", "cuit": "30-71234567-8",
    })
    assert r.status_code == 200, r.text
    assert r.json()["nombre"] == "Compulibra SRL"


def test_el_alta_de_un_cliente_sigue_quedando_en_el_log_de_actividad(client):
    """🔴 La garantia de no-regresion de la Fase 2.

    Al adoptar la tabla `clients` del motor, la tentacion era delegarle el CRUD
    entero a `libracore.db.clients`. No se hizo: ese modulo escribe por su
    conexion DB-API cruda, y el log de actividad de `libraauth` cuelga de los
    eventos de `flush` de la sesion SQLAlchemy. Delegar habria dejado alta,
    edicion y baja de clientes **sin auditar, y sin que nada fallara**.

    Este test es lo que impide que alguien haga esa simplificacion mas
    adelante: si las escrituras se mudan a la conexion cruda, se pone rojo.
    """
    _login(client)
    cid = client.post("/api/clientes", json={
        "nombre": "Auditado", "email": "aud@t.com",
    }).json()["id"]
    client.put(f"/api/clientes/{cid}", json={"nombre": "Auditado SA", "email": "aud@t.com"})

    logs = client.get("/api/logs", params={"entidad": "cliente"}).json()["actividad"]
    acciones = {f["accion"] for f in logs}
    assert "crear" in acciones, f"el alta no quedo registrada: {logs}"
    assert "editar" in acciones, f"la edicion no quedo registrada: {logs}"

    edicion = [f for f in logs if f["accion"] == "editar"][0]
    assert edicion["cambios"] == {"nombre": ["Auditado", "Auditado SA"]}


def _columnas(conn, tabla: str) -> set[str]:
    """Los nombres de columna de una tabla, en el motor que sea.

    Antes esto era `PRAGMA table_info(...)`, que es introspección exclusiva de
    SQLite: contra PostgreSQL el test moría con un error de sintaxis. La suite
    corre contra los dos motores (ver `tests/conftest.py`), así que va por el
    inspector de SQLAlchemy.
    """
    from sqlalchemy import inspect

    return {c["name"] for c in inspect(conn).get_columns(tabla)}


def test_la_migracion_agrega_cuit_y_domicilio_a_una_base_vieja(client):
    """Mismo caso real que la columna de `equipos_movimientos`: los 9 clientes
    de `compulibra` existen desde la migracion del Node.js, y el schema propio
    ya no lo crea `create_all()` sino la cadena de Alembic.

    La base se lleva al baseline con un `downgrade` real, no con `DROP COLUMN` a
    mano: asi el estado de partida es exactamente el que produce la cadena, y de
    paso se ejercita el camino inverso de la revision. Ver tests/test_alembic.py
    para la cobertura del mecanismo; esto verifica el efecto por la API.
    """
    from alembic import command
    from sqlalchemy import text

    from app import database
    from app.schema import BASELINE, _config, ensure_schema

    _login(client)
    cliente_id = client.post("/api/clientes", json={
        "nombre": "Viejo", "email": "v@t.com", "ciudad": "Suipacha",
    }).json()["id"]
    engine = database.get_engine()

    with engine.begin() as conn:
        command.downgrade(_config(conn), BASELINE)
        columnas = _columnas(conn, "clientes")
        filas_antes = conn.execute(text("SELECT COUNT(*) FROM clientes")).scalar()
    assert "cuit" not in columnas and "domicilio" not in columnas
    # El downgrade recrea `clientes` en batch: la fila tiene que sobrevivir.
    assert filas_antes == 1

    assert ensure_schema(engine) == "upgrade"

    with engine.begin() as conn:
        columnas = _columnas(conn, "clientes")
        filas_despues = conn.execute(text("SELECT COUNT(*) FROM clientes")).scalar()
    assert {"cuit", "domicilio"} <= columnas
    # El cliente migrado no se pierde: queda con los campos nuevos en NULL.
    assert filas_despues == filas_antes
    ficha = client.get(f"/api/clientes/{cliente_id}").json()
    assert ficha["ciudad"] == "Suipacha" and ficha["cuit"] is None

    assert ensure_schema(engine) == "upgrade"  # idempotente

    # Y la base migrada acepta escribir las columnas nuevas, que es el punto.
    r = client.put(f"/api/clientes/{cliente_id}", json={
        "nombre": "Viejo", "email": "v@t.com", "cuit": "20-11111111-2",
    })
    assert r.json()["cuit"] == "20-11111111-2"


def test_incidencias_requires_auth(client):
    r = client.get("/api/incidencias")
    assert r.status_code == 401


def test_usuarios_requires_admin(client):
    _login(client)
    r = client.get("/api/usuarios")
    assert r.status_code == 200  # admin logueado


def test_usuario_duplicado_devuelve_409(client):
    """Bug latente que estuvo vivo desde el dia 1 y no tenia test.

    Un username repetido tiene que dar 409, no 500. Con el pin viejo de
    libraauth (`v0.1.0`) el duplicado propagaba el `IntegrityError` de
    SQLAlchemy; desde `v0.1.1` levanta `UsernameTaken`, que **no hereda de
    ValueError**, asi que el `except ValueError` del router tampoco lo
    agarraba: bumpear el pin solo no arreglaba nada. Ver
    wiki/entities/libraauth.md.
    """
    _login(client)
    alta = {"username": "repetido", "name": "Repetido", "password": "secreta123",
            "role": "admin"}
    assert client.post("/api/usuarios", json=alta).status_code == 201
    assert client.post("/api/usuarios", json=alta).status_code == 409


def test_usuario_con_rol_invalido_devuelve_422(client):
    """El otro camino del mismo `try`: que el `except UsernameTaken` nuevo no
    se coma el 422 que ya existia."""
    _login(client)
    r = client.post("/api/usuarios", json={
        "username": "rolmalo", "name": "Rol Malo", "password": "secreta123",
        "role": "rol-que-no-existe"})
    assert r.status_code == 422


def test_export_xlsx(client):
    _login(client)
    client.post("/api/clientes", json={"nombre": "Cliente XLSX"})
    r = client.get("/api/reportes/clientes.xlsx")
    assert r.status_code == 200
    assert r.headers["content-type"] == XLSX_MIME


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _equipo_de_prueba(client, **extra) -> int:
    cliente_id = client.post("/api/clientes", json={"nombre": "Cliente Mov"}).json()["id"]
    body = {"cliente_id": cliente_id, "tipo": "Notebook", "marca": "Lenovo",
            "sector": "Administracion", "ubicacion_oficina": "Piso 2",
            "estado": "activo"}
    body.update(extra)
    return client.post("/api/equipos", json=body).json()["id"]


def test_alta_de_equipo_registra_movimiento(client):
    _login(client)
    equipo_id = _equipo_de_prueba(client)

    movs = client.get(f"/api/equipos/{equipo_id}/movimientos").json()
    assert len(movs) == 1
    assert movs[0]["tipo"] == "alta"
    assert movs[0]["descripcion"] == "Alta: Notebook Lenovo"
    assert movs[0]["sector_destino"] == "Administracion"
    assert movs[0]["ubicacion_destino"] == "Piso 2"
    # El actor queda registrado, no "Sistema".
    assert movs[0]["usuario"] == "admin"


def test_traslado_registra_origen_y_destino(client):
    _login(client)
    equipo_id = _equipo_de_prueba(client)

    client.put(f"/api/equipos/{equipo_id}", json={
        "cliente_id": 1, "tipo": "Notebook", "marca": "Lenovo",
        "sector": "Deposito", "ubicacion_oficina": "Subsuelo",
        "estado": "activo", "motivo": "Reubicacion por obra",
    })

    movs = client.get(f"/api/equipos/{equipo_id}/movimientos").json()
    traslado = next(m for m in movs if m["tipo"] == "traslado")
    assert traslado["sector_origen"] == "Administracion"
    assert traslado["sector_destino"] == "Deposito"
    assert traslado["ubicacion_origen"] == "Piso 2"
    assert traslado["ubicacion_destino"] == "Subsuelo"
    assert traslado["motivo"] == "Reubicacion por obra"
    assert traslado["descripcion"] == "Traslado → Deposito"


def test_cambio_de_estado_registra_movimiento_con_ese_tipo(client):
    """El `tipo` del movimiento es el estado nuevo: asi el reporte lo
    etiqueta 'Reparación'/'Baja'/'Reactivado' via MOV_LABEL."""
    _login(client)
    equipo_id = _equipo_de_prueba(client)

    base = {"cliente_id": 1, "tipo": "Notebook", "marca": "Lenovo",
            "sector": "Administracion", "ubicacion_oficina": "Piso 2"}
    client.put(f"/api/equipos/{equipo_id}", json={**base, "estado": "en_reparacion",
                                                 "motivo": "No enciende"})

    movs = client.get(f"/api/equipos/{equipo_id}/movimientos").json()
    cambio = next(m for m in movs if m["tipo"] == "en_reparacion")
    assert cambio["descripcion"] == "Estado cambiado a: en_reparacion"
    assert cambio["motivo"] == "No enciende"


def test_un_update_sin_cambios_relevantes_no_ensucia_el_historial(client):
    """Corregir el serial no es un movimiento: si cada PUT generara una
    fila, el historial se volveria inservible."""
    _login(client)
    equipo_id = _equipo_de_prueba(client)
    antes = len(client.get(f"/api/equipos/{equipo_id}/movimientos").json())

    client.put(f"/api/equipos/{equipo_id}", json={
        "cliente_id": 1, "tipo": "Notebook", "marca": "Lenovo",
        "sector": "Administracion", "ubicacion_oficina": "Piso 2",
        "estado": "activo", "serial": "CORREGIDO-123",
    })

    despues = client.get(f"/api/equipos/{equipo_id}/movimientos").json()
    assert len(despues) == antes


def test_traslado_y_cambio_de_estado_juntos_generan_dos_movimientos(client):
    """Un equipo que vuelve del service Y cambia de sector son dos hechos
    distintos, y el historial tiene que reflejar los dos."""
    _login(client)
    equipo_id = _equipo_de_prueba(client, estado="en_reparacion")

    client.put(f"/api/equipos/{equipo_id}", json={
        "cliente_id": 1, "tipo": "Notebook", "marca": "Lenovo",
        "sector": "Consultorios", "ubicacion_oficina": "Consultorio 4",
        "estado": "activo",
    })

    movs = client.get(f"/api/equipos/{equipo_id}/movimientos").json()
    tipos = [m["tipo"] for m in movs]
    assert "traslado" in tipos
    assert "activo" in tipos
    assert len(movs) == 3  # alta + traslado + reactivacion


# ─── Trazabilidad incidencia ↔ activo ───────────────────────────────────

def _escenario_impresora(client) -> dict:
    """El caso real que motivo todo esto: una impresora que hace ruido en
    Admision, y una de repuesto en el deposito."""
    cliente_id = client.post("/api/clientes", json={"nombre": "Hospital"}).json()["id"]
    hp = client.post("/api/equipos", json={
        "cliente_id": cliente_id, "tipo": "Impresora", "marca": "HP",
        "modelo": "LaserJet M501", "sector": "Admision", "ubicacion_oficina": "Mostrador",
        "estado": "activo",
    }).json()["id"]
    pantum = client.post("/api/equipos", json={
        "cliente_id": cliente_id, "tipo": "Impresora", "marca": "Pantum",
        "modelo": "M6559", "sector": "Deposito", "estado": "almacenado",
    }).json()["id"]
    incidencia_id = client.post("/api/incidencias", json={
        "cliente_id": cliente_id, "equipo_id": hp,
        "titulo": "Ruido anormal al imprimir", "prioridad": "media",
    }).json()["id"]
    return {"cliente_id": cliente_id, "hp": hp, "pantum": pantum,
            "incidencia_id": incidencia_id}


def test_incidencias_se_filtran_por_equipo(client):
    """'¿Cuántas veces falló esta impresora?' — el dato estaba desde la
    migracion pero no habia forma de pedirlo: el listado solo filtraba por
    cliente y estado."""
    _login(client)
    esc = _escenario_impresora(client)

    client.post("/api/incidencias", json={
        "cliente_id": esc["cliente_id"], "equipo_id": esc["hp"], "titulo": "Segunda falla",
    })
    client.post("/api/incidencias", json={
        "cliente_id": esc["cliente_id"], "equipo_id": esc["pantum"], "titulo": "Otra impresora",
    })

    de_la_hp = client.get(f"/api/incidencias?equipo_id={esc['hp']}").json()
    assert len(de_la_hp) == 2
    assert {i["titulo"] for i in de_la_hp} == {"Ruido anormal al imprimir", "Segunda falla"}
    # Y sigue combinando con los filtros que ya existian.
    assert len(client.get(
        f"/api/incidencias?equipo_id={esc['hp']}&estado=cerrado").json()) == 0


def test_reemplazo_mueve_los_dos_equipos_y_deja_todo_ligado(client):
    """La operacion completa en una llamada: la HP sale a service, la
    Pantum ocupa su lugar, los 4 movimientos quedan atados al ticket y las
    dos intervenciones narradas."""
    _login(client)
    esc = _escenario_impresora(client)

    r = client.post(f"/api/incidencias/{esc['incidencia_id']}/reemplazar-equipo", json={
        "equipo_retirado_id": esc["hp"],
        "equipo_sustituto_id": esc["pantum"],
        "destino": "service",
        "motivo": "Ruido mecanico en el conjunto de rodillos",
    })
    assert r.status_code == 201, r.text
    resultado = r.json()

    # El equipo retirado quedo en service…
    assert resultado["retirado"]["estado"] == "en_reparacion"
    assert resultado["retirado"]["sector"] == "Service"
    # …y el sustituto ocupa EXACTAMENTE el lugar que dejo.
    assert resultado["sustituto"]["estado"] == "activo"
    assert resultado["sustituto"]["sector"] == "Admision"
    assert resultado["sustituto"]["ubicacion_oficina"] == "Mostrador"

    # Los 4 movimientos (traslado + estado, por equipo) traen el ticket.
    movs = client.get(f"/api/incidencias/{esc['incidencia_id']}/movimientos").json()
    assert len(movs) == 4
    assert all(m["incidencia_id"] == esc["incidencia_id"] for m in movs)
    assert {m["equipo_id"] for m in movs} == {esc["hp"], esc["pantum"]}

    # Y la incidencia cuenta la historia sola, sin que nadie escriba nada.
    actividades = client.get(f"/api/incidencias/{esc['incidencia_id']}/actividades").json()
    textos = " ".join(a["descripcion"] for a in actividades)
    assert "Se retira Impresora HP LaserJet M501 de Admision · Mostrador" in textos
    assert "se envía a service" in textos
    assert "Se instala Impresora Pantum M6559 en Admision · Mostrador" in textos
    assert "Ruido mecanico" in textos


def test_el_reemplazo_queda_en_orden_cronologico_real(client):
    """Defecto encontrado **probando la UI, no con los tests**: las 6 filas
    de un reemplazo caian en el mismo segundo (`CURRENT_TIMESTAMP` de
    SQLite no tiene fraccion), y el timeline mostraba la instalacion del
    sustituto ANTES del retiro del equipo al que venia a reemplazar.

    Las fechas ahora se sellan con microsegundos en orden causal: primero
    el retiro y sus movimientos, despues la instalacion y los suyos.
    """
    _login(client)
    esc = _escenario_impresora(client)
    inc = esc["incidencia_id"]

    client.post(f"/api/incidencias/{inc}/reemplazar-equipo", json={
        "equipo_retirado_id": esc["hp"], "equipo_sustituto_id": esc["pantum"],
        "destino": "service",
    })

    actividades = client.get(f"/api/incidencias/{inc}/actividades").json()
    movimientos = client.get(f"/api/incidencias/{inc}/movimientos").json()
    todo = sorted([*actividades, *movimientos], key=lambda f: f["fecha"])

    # Ninguna fecha empatada **con resolucion de milisegundo**, que es la
    # del `Date` de JavaScript: sellarlas con microsegundos las deja
    # distintas en la base y empatadas en el navegador — probado, el
    # timeline seguia desordenado.
    fechas_ms = [f["fecha"][:23] for f in todo]
    assert len(set(fechas_ms)) == len(fechas_ms), fechas_ms

    # Y la historia se lee en el orden en que paso.
    def es_del(fila, equipo_id):
        return fila.get("equipo_id") == equipo_id

    assert "Se retira" in todo[0]["descripcion"]
    assert all(es_del(f, esc["hp"]) for f in todo[1:3])
    assert "Se instala" in todo[3]["descripcion"]
    assert all(es_del(f, esc["pantum"]) for f in todo[4:6])


def test_la_vuelta_del_service_es_la_misma_operacion_al_reves(client):
    """No hace falta una accion aparte para reinstalar: se reemplaza la
    prestada por la que volvio, con destino deposito."""
    _login(client)
    esc = _escenario_impresora(client)
    inc = esc["incidencia_id"]

    client.post(f"/api/incidencias/{inc}/reemplazar-equipo", json={
        "equipo_retirado_id": esc["hp"], "equipo_sustituto_id": esc["pantum"],
        "destino": "service",
    })
    r = client.post(f"/api/incidencias/{inc}/reemplazar-equipo", json={
        "equipo_retirado_id": esc["pantum"], "equipo_sustituto_id": esc["hp"],
        "destino": "deposito", "motivo": "Reparada, se reinstala",
    })
    assert r.status_code == 201, r.text

    hp = client.get(f"/api/equipos/{esc['hp']}").json()
    pantum = client.get(f"/api/equipos/{esc['pantum']}").json()
    assert (hp["estado"], hp["sector"], hp["ubicacion_oficina"]) == \
        ("activo", "Admision", "Mostrador")
    assert (pantum["estado"], pantum["sector"]) == ("almacenado", "Depósito")

    # El historial de la HP cuenta el viaje completo, en orden inverso.
    tipos = [m["tipo"] for m in client.get(f"/api/equipos/{esc['hp']}/movimientos").json()]
    assert tipos == ["activo", "traslado", "en_reparacion", "traslado", "alta"]


def test_reemplazo_sin_sustituto_es_valido(client):
    """No siempre hay repuesto a mano: retirar sin reponer tiene que
    poder registrarse, en vez de obligar a inventar un sustituto."""
    _login(client)
    esc = _escenario_impresora(client)

    r = client.post(f"/api/incidencias/{esc['incidencia_id']}/reemplazar-equipo", json={
        "equipo_retirado_id": esc["hp"], "destino": "baja",
    })
    assert r.status_code == 201, r.text
    assert r.json()["sustituto"] is None
    assert r.json()["retirado"]["estado"] == "baja"
    assert len(r.json()["actividades"]) == 1
    # La Pantum no se movio de su lugar.
    assert client.get(f"/api/equipos/{esc['pantum']}").json()["sector"] == "Deposito"


def test_reemplazo_rechaza_los_casos_imposibles(client):
    _login(client)
    esc = _escenario_impresora(client)
    inc = esc["incidencia_id"]
    url = f"/api/incidencias/{inc}/reemplazar-equipo"

    # El mismo equipo de los dos lados.
    assert client.post(url, json={"equipo_retirado_id": esc["hp"],
                                  "equipo_sustituto_id": esc["hp"]}).status_code == 422
    # Destino que no existe.
    assert client.post(url, json={"equipo_retirado_id": esc["hp"],
                                  "destino": "marte"}).status_code == 422
    # Equipo inexistente e incidencia inexistente.
    assert client.post(url, json={"equipo_retirado_id": 99999}).status_code == 404
    assert client.post("/api/incidencias/99999/reemplazar-equipo",
                       json={"equipo_retirado_id": esc["hp"]}).status_code == 404
    # Un ticket de un cliente no puede retirar el equipo de otro.
    otro = client.post("/api/clientes", json={"nombre": "Otro"}).json()["id"]
    ajeno = client.post("/api/equipos", json={"cliente_id": otro, "tipo": "PC"}).json()["id"]
    assert client.post(url, json={"equipo_retirado_id": ajeno}).status_code == 422


def test_un_reemplazo_fallido_no_deja_nada_a_medias(client, monkeypatch):
    """La atomicidad es el punto de la accion compuesta: media operacion
    aplicada es peor que ninguna — el inventario diria que la impresora
    esta en service y el ticket no tendria ni una linea que lo explique.

    Se **fuerza el fallo en el medio** (despues de mover el retirado,
    antes de terminar) en vez de confiar en que las validaciones de
    entrada alcanzan: esas corren todas antes de la primera escritura, asi
    que un `equipo_sustituto_id` inexistente no probaria nada sobre la
    transaccion.
    """
    _login(client)
    esc = _escenario_impresora(client)

    from app.services import reemplazo as modulo

    original = modulo.movimientos_por_cambio
    llamadas = {"n": 0}

    def explota_en_el_segundo_equipo(*args, **kwargs):
        llamadas["n"] += 1
        if llamadas["n"] == 2:
            raise RuntimeError("fallo simulado a mitad de camino")
        return original(*args, **kwargs)

    monkeypatch.setattr(modulo, "movimientos_por_cambio", explota_en_el_segundo_equipo)

    antes = client.get(f"/api/equipos/{esc['hp']}").json()
    with pytest.raises(RuntimeError):
        client.post(f"/api/incidencias/{esc['incidencia_id']}/reemplazar-equipo", json={
            "equipo_retirado_id": esc["hp"], "equipo_sustituto_id": esc["pantum"],
        })

    assert llamadas["n"] == 2  # el fallo ocurrio DESPUES de tocar el primer equipo
    assert client.get(f"/api/equipos/{esc['hp']}").json() == antes
    assert client.get(f"/api/equipos/{esc['pantum']}").json()["sector"] == "Deposito"
    assert client.get(f"/api/incidencias/{esc['incidencia_id']}/movimientos").json() == []
    assert client.get(f"/api/incidencias/{esc['incidencia_id']}/actividades").json() == []


def test_un_sustituto_inexistente_da_404_antes_de_tocar_nada(client):
    """Las validaciones corren antes de la primera escritura."""
    _login(client)
    esc = _escenario_impresora(client)

    antes = client.get(f"/api/equipos/{esc['hp']}").json()
    r = client.post(f"/api/incidencias/{esc['incidencia_id']}/reemplazar-equipo", json={
        "equipo_retirado_id": esc["hp"], "equipo_sustituto_id": 99999,
    })
    assert r.status_code == 404
    assert client.get(f"/api/equipos/{esc['hp']}").json() == antes


def test_borrar_la_incidencia_no_borra_el_movimiento_pero_lo_desvincula(client):
    """El equipo salio de Admision de verdad: ese hecho fisico sobrevive
    al ticket. Y como el pragma foreign_keys esta apagado en las
    conexiones de SQLAlchemy, esto se hace explicito en el repositorio —
    si se confiara en el `ondelete`, el movimiento quedaria apuntando a un
    ticket borrado."""
    _login(client)
    esc = _escenario_impresora(client)
    client.post(f"/api/incidencias/{esc['incidencia_id']}/reemplazar-equipo", json={
        "equipo_retirado_id": esc["hp"], "destino": "service",
    })

    assert client.delete(f"/api/incidencias/{esc['incidencia_id']}").status_code == 204

    movs = client.get(f"/api/equipos/{esc['hp']}/movimientos").json()
    assert len(movs) == 3  # alta + traslado + en_reparacion, intactos
    assert all(m["incidencia_id"] is None for m in movs)


def test_borrar_la_incidencia_borra_su_actividad_y_su_auditoria(client):
    """Lo que el dialogo de confirmacion de la UI ya prometia y no pasaba:
    los `ondelete=CASCADE` de los modelos son decorativos porque el engine
    no activa `PRAGMA foreign_keys` (medido: devuelve 0)."""
    _login(client)
    esc = _escenario_impresora(client)
    inc = esc["incidencia_id"]
    client.post(f"/api/incidencias/{inc}/actividades", json={"descripcion": "Diagnostico"})

    from sqlalchemy import text

    from app import database

    def contar(tabla: str) -> int:
        with database.get_engine().connect() as conn:
            return conn.execute(
                text(f"SELECT COUNT(*) FROM {tabla} WHERE incidencia_id = {inc}")
            ).scalar()

    assert contar("actividades_incidencia") == 1
    assert contar("incidencias_estados_log") == 1

    assert client.delete(f"/api/incidencias/{inc}").status_code == 204

    assert contar("actividades_incidencia") == 0
    assert contar("incidencias_estados_log") == 0


# El test que cubria `app/migrations.py` (el `ALTER TABLE ADD COLUMN` a mano)
# se fue con ese modulo el 2026-08-03, cuando el schema propio paso a Alembic.
# Su sucesor vive en tests/test_alembic.py: alli se reconstruye el schema real
# de produccion —el mismo `CREATE TABLE` que estaba escrito aca— y se verifica
# que `ensure_schema()` lo adopte sin tocar el schema ni perder filas.




def _armar_datos_para_reportes(client) -> dict:
    """Un cliente por_servicio con equipo (garantia vencida), incidencia
    cerrada, actividad y movimiento — toca las 6 consultas analiticas."""
    cliente_id = client.post("/api/clientes", json={
        "nombre": "Reportes SA", "empresa": "Reportes SA",
        "tipo_facturacion": "por_servicio",
    }).json()["id"]
    equipo_id = client.post("/api/equipos", json={
        "cliente_id": cliente_id, "tipo": "Notebook", "marca": "Lenovo",
        "garantia_vence": "2020-01-01", "estado": "activo",
    }).json()["id"]
    tecnico_id = client.post("/api/tecnicos", json={"nombre": "Tec Reportes"}).json()["id"]
    incidencia_id = client.post("/api/incidencias", json={
        "cliente_id": cliente_id, "equipo_id": equipo_id, "tecnico_id": tecnico_id,
        "titulo": "Falla de fuente", "prioridad": "alta",
    }).json()["id"]
    client.post(f"/api/incidencias/{incidencia_id}/actividades",
                json={"descripcion": "Diagnostico inicial"})
    client.put(f"/api/incidencias/{incidencia_id}", json={
        "cliente_id": cliente_id, "equipo_id": equipo_id, "tecnico_id": tecnico_id,
        "titulo": "Falla de fuente", "prioridad": "alta", "estado": "cerrado",
    })
    return {"cliente_id": cliente_id, "equipo_id": equipo_id}


def test_reportes_analiticos_devuelven_xlsx(client):
    """Los 6 reportes reconstruidos responden 200 con un xlsx real. Se
    verifica la firma ZIP ('PK') y no solo el content-type: un 200 con
    cuerpo vacio pasaria el chequeo de header igual."""
    _login(client)
    _armar_datos_para_reportes(client)
    periodo = "desde=2020-01-01&hasta=2030-12-31"

    for url in [
        "/api/reportes/equipamiento.xlsx",
        f"/api/reportes/incidencias-periodo.xlsx?{periodo}",
        f"/api/reportes/facturacion.xlsx?{periodo}",
        "/api/reportes/garantias.xlsx?dias=60",
        f"/api/reportes/tecnico.xlsx?{periodo}",
        f"/api/reportes/movimientos.xlsx?{periodo}",
    ]:
        r = client.get(url)
        assert r.status_code == 200, f"{url} -> {r.status_code}"
        assert r.headers["content-type"] == XLSX_MIME, url
        assert r.content[:2] == b"PK", f"{url} no devolvio un xlsx real"


def test_reporte_periodo_exige_fechas(client):
    """`desde`/`hasta` son obligatorios: sin ellos el reporte no tiene
    sentido y el original respondia 400."""
    _login(client)
    assert client.get("/api/reportes/incidencias-periodo.xlsx").status_code == 422


def test_reportes_exigen_autenticacion(client):
    assert client.get("/api/reportes/garantias.xlsx").status_code == 401


def test_contenido_real_del_reporte_por_tecnico(client):
    """Lee el xlsx con openpyxl y confirma los numeros, no solo que baje:
    un reporte que devuelve un archivo vacio tambien daria 200."""
    from io import BytesIO

    from openpyxl import load_workbook

    _login(client)
    _armar_datos_para_reportes(client)
    r = client.get("/api/reportes/tecnico.xlsx?desde=2020-01-01&hasta=2030-12-31")
    ws = load_workbook(BytesIO(r.content)).active

    # Fila 4 = headers; 5 = primer tecnico; ultima = totales.
    assert ws.cell(row=4, column=1).value == "Técnico"
    assert ws.cell(row=5, column=1).value == "Tec Reportes"
    assert ws.cell(row=5, column=2).value == 1  # total
    assert ws.cell(row=5, column=5).value == 1  # cerradas
    assert ws.cell(row=5, column=6).value == "100%"  # % resolucion
    assert ws.cell(row=5, column=7).value == 1  # actividades
    assert ws.cell(row=6, column=1).value == "TOTAL"


def test_garantia_vencida_se_reporta_como_vencida(client):
    """El equipo del fixture vencio en 2020: la columna de dias tiene que
    decir 'Vencida hace Xd', no un numero de dias positivo."""
    from io import BytesIO

    from openpyxl import load_workbook

    _login(client)
    _armar_datos_para_reportes(client)
    r = client.get("/api/reportes/garantias.xlsx?dias=60")
    ws = load_workbook(BytesIO(r.content)).active
    assert str(ws.cell(row=5, column=9).value).startswith("Vencida hace")


def test_facturacion_solo_incluye_clientes_por_servicio(client):
    """Un cliente 'mensual' cobra abono, no incidencia: sus cerradas no
    deben aparecer en el reporte de facturacion."""
    from io import BytesIO

    from openpyxl import load_workbook

    _login(client)
    _armar_datos_para_reportes(client)  # por_servicio -> si aparece

    mensual_id = client.post("/api/clientes", json={
        "nombre": "Abono SA", "tipo_facturacion": "mensual",
    }).json()["id"]
    inc = client.post("/api/incidencias", json={
        "cliente_id": mensual_id, "titulo": "Incidencia de abono",
    }).json()["id"]
    client.put(f"/api/incidencias/{inc}", json={
        "cliente_id": mensual_id, "titulo": "Incidencia de abono", "estado": "cerrado",
    })

    r = client.get("/api/reportes/facturacion.xlsx?desde=2020-01-01&hasta=2030-12-31")
    ws = load_workbook(BytesIO(r.content)).active
    textos = [
        str(c.value) for fila in ws.iter_rows() for c in fila if c.value is not None
    ]
    assert any("Falla de fuente" in t for t in textos)
    assert not any("Incidencia de abono" in t for t in textos)


# ─── Remitos y presupuestos (dominio reusado de libracore) ──────────────

def _cliente_para_comprobantes(client) -> int:
    return client.post("/api/clientes", json={
        "nombre": "Juan Perez", "empresa": "Compulibra SRL",
        "email": "facturacion@compulibra.com.ar", "telefono": "3514567890",
        "ciudad": "Cordoba",
    }).json()["id"]


_ITEMS = [
    {"description": "Reparacion de notebook", "qty": 2, "unit_price": 15000},
    {"description": "Cambio de disco SSD 480GB", "qty": 1, "unit_price": 45000},
]


def test_remito_con_client_id_real_se_inserta(client):
    """El test que justifica el diseno del schema: el DDL de LibraCore declara
    `client_id REFERENCES clients(id)` y LibraDesk no tiene `clients`. Como
    `libracore.db.core.get_connection()` corre `PRAGMA foreign_keys = ON` en
    toda conexion, dejar esa FK haria fallar TODO insert con `no such table:
    main.clients` — incluso con client_id NULL. Si alguien "restaura" la FK
    copiando el DDL original, este test se pone rojo."""
    _login(client)
    cliente_id = _cliente_para_comprobantes(client)

    r = client.post("/api/remitos", json={"client_id": cliente_id, "items": _ITEMS})
    assert r.status_code == 201, r.text
    remito = r.json()

    assert remito["client_id"] == cliente_id
    assert remito["client_name"] == "Compulibra SRL"   # empresa, no nombre
    assert remito["client_email"] == "facturacion@compulibra.com.ar"
    assert remito["client_address"] == "Cordoba"
    assert remito["number"].startswith("0001-")
    assert len(remito["items"]) == 2


def test_remito_calcula_los_totales_en_el_servidor(client):
    """Los importes no se aceptan del front: 2*15000 + 1*45000 = 75000."""
    _login(client)
    cliente_id = _cliente_para_comprobantes(client)

    remito = client.post("/api/remitos", json={
        "client_id": cliente_id, "items": _ITEMS, "tax_rate": 0.21,
    }).json()

    assert remito["subtotal"] == 75000
    assert remito["tax_amount"] == 15750
    assert remito["total"] == 90750
    assert remito["items"][0]["subtotal"] == 30000


def test_remito_de_cliente_inexistente_da_404(client):
    """La FK no existe, asi que la integridad la sostiene el router."""
    _login(client)
    r = client.post("/api/remitos", json={"client_id": 99999, "items": _ITEMS})
    assert r.status_code == 404


def test_remito_sin_items_es_rechazado(client):
    _login(client)
    cliente_id = _cliente_para_comprobantes(client)
    r = client.post("/api/remitos", json={"client_id": cliente_id, "items": []})
    assert r.status_code == 422


def test_next_number_no_choca_con_la_ruta_de_detalle(client):
    """/next-number va declarada antes de /{remito_id}; al reves daria 422
    porque "next-number" no parsea como int."""
    _login(client)
    r = client.get("/api/remitos/next-number")
    assert r.status_code == 200
    assert r.json()["number"].startswith("0001-")


def test_remitos_exigen_autenticacion(client):
    """Los routers de dominio ya se habian olvidado la autenticacion una vez
    (solo /api/usuarios la tenia); se cubre explicitamente."""
    assert client.get("/api/remitos").status_code == 401
    assert client.get("/api/presupuestos").status_code == 401
    assert client.post("/api/remitos", json={"client_id": 1, "items": _ITEMS}).status_code == 401


def test_remito_pdf_es_un_pdf_real(client):
    """No alcanza con un 200: se verifica la firma del archivo y que el
    generador de LibraCore produjo algo con contenido."""
    _login(client)
    cliente_id = _cliente_para_comprobantes(client)
    remito_id = client.post("/api/remitos", json={
        "client_id": cliente_id, "items": _ITEMS, "observations": "Retira el cliente",
    }).json()["id"]

    r = client.get(f"/api/remitos/{remito_id}/pdf")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content.startswith(b"%PDF")
    assert len(r.content) > 1000

    # El path queda guardado en la fila, como en Contalibra/Restolibra.
    assert client.get(f"/api/remitos/{remito_id}").json()["pdf_path"].endswith(".pdf")


def test_presupuesto_flujo_completo_y_conversion_a_remito(client):
    _login(client)
    cliente_id = _cliente_para_comprobantes(client)

    p = client.post("/api/presupuestos", json={
        "client_id": cliente_id, "items": _ITEMS,
    })
    assert p.status_code == 201, p.text
    presupuesto = p.json()
    assert presupuesto["number"].startswith("PRES-")
    assert presupuesto["status"] == "borrador"
    assert presupuesto["total"] == 90750

    pid = presupuesto["id"]
    r = client.patch(f"/api/presupuestos/{pid}/estado", json={"status": "enviado"})
    assert r.status_code == 200
    assert r.json()["status"] == "enviado"

    r = client.post(f"/api/presupuestos/{pid}/convertir-en-remito")
    assert r.status_code == 201, r.text
    remito = r.json()
    assert remito["total"] == presupuesto["total"]
    assert presupuesto["number"] in remito["observations"]

    despues = client.get(f"/api/presupuestos/{pid}").json()
    assert despues["status"] == "aceptado"
    assert despues["remito_id"] == remito["id"]


def test_convertir_dos_veces_no_emite_un_segundo_remito(client):
    """Idempotente a proposito: dos clicks no facturan el trabajo dos veces."""
    _login(client)
    cliente_id = _cliente_para_comprobantes(client)
    pid = client.post("/api/presupuestos", json={
        "client_id": cliente_id, "items": _ITEMS,
    }).json()["id"]

    primero = client.post(f"/api/presupuestos/{pid}/convertir-en-remito").json()
    segundo = client.post(f"/api/presupuestos/{pid}/convertir-en-remito").json()

    assert primero["id"] == segundo["id"]
    assert len(client.get("/api/remitos").json()) == 1


def test_presupuesto_enviado_y_expirado_aparece_vencido_solo(client):
    """El vencimiento lo corre LibraCore al leer, sin tarea programada: un
    `enviado` con valid_until pasado sale `vencido` del listado."""
    _login(client)
    cliente_id = _cliente_para_comprobantes(client)

    pid = client.post("/api/presupuestos", json={
        "client_id": cliente_id, "items": _ITEMS,
        "valid_until": "2020-01-01", "status": "enviado",
    }).json()["id"]

    # El detalle no dispara el vencimiento; el listado si.
    assert client.get(f"/api/presupuestos/{pid}").json()["status"] == "enviado"
    listado = client.get("/api/presupuestos").json()
    assert next(p for p in listado if p["id"] == pid)["status"] == "vencido"
    assert client.get("/api/presupuestos/resumen").json()["vencido"] == 1


def test_convertir_un_presupuesto_vencido_da_409(client):
    _login(client)
    cliente_id = _cliente_para_comprobantes(client)
    pid = client.post("/api/presupuestos", json={
        "client_id": cliente_id, "items": _ITEMS,
        "valid_until": "2020-01-01", "status": "enviado",
    }).json()["id"]
    client.get("/api/presupuestos")  # dispara el vencimiento

    r = client.post(f"/api/presupuestos/{pid}/convertir-en-remito")
    assert r.status_code == 409


def test_solo_se_borra_un_presupuesto_en_borrador(client):
    """Regla de LibraCore, no propia: en cualquier otro estado levanta
    ValueError, que el router traduce a 409."""
    _login(client)
    cliente_id = _cliente_para_comprobantes(client)

    borrador = client.post("/api/presupuestos", json={
        "client_id": cliente_id, "items": _ITEMS,
    }).json()["id"]
    assert client.delete(f"/api/presupuestos/{borrador}").status_code == 204

    enviado = client.post("/api/presupuestos", json={
        "client_id": cliente_id, "items": _ITEMS, "status": "enviado",
    }).json()["id"]
    assert client.delete(f"/api/presupuestos/{enviado}").status_code == 409
    assert client.get(f"/api/presupuestos/{enviado}").status_code == 200


def test_estado_invalido_es_rechazado(client):
    _login(client)
    cliente_id = _cliente_para_comprobantes(client)
    pid = client.post("/api/presupuestos", json={
        "client_id": cliente_id, "items": _ITEMS,
    }).json()["id"]
    assert client.patch(f"/api/presupuestos/{pid}/estado",
                        json={"status": "pagado"}).status_code == 422


def test_presupuesto_pdf_es_un_pdf_real(client):
    _login(client)
    cliente_id = _cliente_para_comprobantes(client)
    pid = client.post("/api/presupuestos", json={
        "client_id": cliente_id, "items": _ITEMS,
    }).json()["id"]

    r = client.get(f"/api/presupuestos/{pid}/pdf")
    assert r.status_code == 200
    assert r.content.startswith(b"%PDF")
    assert len(r.content) > 1000


def test_el_pdf_no_se_cae_con_un_guion_largo(client):
    """Regresión del 500 del 2026-08-03, y **guarda del pin de libracore**.

    En `libradesk-dev` el nombre de empresa era ``"Compulibra — Soporte IT"`` y
    **todo** PDF de presupuesto devolvía 500: las fuentes core de fpdf2 se
    codificaban en latin-1, donde el guión largo no entra. El arreglo está en
    libracore >= v1.7.0.

    Los dos tests de PDF de acá arriba no lo agarraban porque usan texto ASCII.
    Éste vive en este repo y no sólo en libracore a propósito: lo que puede
    volver a romperse desde acá es **el pin**, y un pin viejo lo pone en rojo.

    Se cubren los tres orígenes del texto —empresa, ítem y observaciones—
    porque los tres llegan al PDF por caminos distintos.
    """
    _login(client)
    client.put("/api/config/empresa", json={"empresa_nombre": "Compulibra — Soporte IT"})
    cliente_id = _cliente_para_comprobantes(client)
    pid = client.post("/api/presupuestos", json={
        "client_id": cliente_id,
        "observations": "Reparación “urgente” — entrega en 48 h…",
        "items": [{"description": "Mano de obra — diagnóstico", "qty": 1, "unit_price": 1000}],
    }).json()["id"]

    r = client.get(f"/api/presupuestos/{pid}/pdf")
    assert r.status_code == 200, r.text
    assert r.content.startswith(b"%PDF")

    # Y el remito, que arma el encabezado por el mismo camino.
    rid = client.post("/api/remitos", json={
        "client_id": cliente_id,
        "items": [{"description": "Notebook — reemplazo", "qty": 1, "unit_price": 1000}],
    }).json()["id"]
    assert client.get(f"/api/remitos/{rid}/pdf").status_code == 200


def test_config_empresa_ida_y_vuelta(client):
    """El encabezado de los PDF. Sin esto salen con la empresa en blanco."""
    _login(client)
    assert client.get("/api/config/empresa").json()["empresa_nombre"] == ""

    r = client.put("/api/config/empresa", json={
        "empresa_nombre": "Compulibra", "empresa_cuit": "20-12345678-9",
        "empresa_direccion": "Suipacha 123", "empresa_email": "info@compulibra.com.ar",
    })
    assert r.status_code == 200
    assert r.json()["empresa_nombre"] == "Compulibra"
    assert client.get("/api/config/empresa").json()["empresa_cuit"] == "20-12345678-9"


def test_config_empresa_no_expone_ni_pisa_secretos(client):
    """config.json es compartido con MercadoPago/SMTP en la familia: esta API
    solo toca las claves empresa_*, y no devuelve las otras."""
    from libracore import config_manager
    _login(client)

    cfg = config_manager.load()
    cfg["mp_access_token"] = "TOKEN-QUE-NO-SE-DEBE-PERDER"
    config_manager.save(cfg)

    r = client.get("/api/config/empresa")
    assert "mp_access_token" not in r.json()

    client.put("/api/config/empresa", json={"empresa_nombre": "Compulibra"})
    assert config_manager.load()["mp_access_token"] == "TOKEN-QUE-NO-SE-DEBE-PERDER"


# --- Sectores -------------------------------------------------------------
#
# El router existia desde la migracion desde el Node.js viejo (con 15 filas
# reales en produccion) y no tenia un solo test, porque hasta el 2026-07-31
# no habia ninguna pantalla que lo llamara: se llegaba nada mas que desde el
# selector de filtro de un reporte.


def test_sectores_crud_por_cliente(client):
    _login(client)
    a = client.post("/api/clientes", json={"nombre": "Cliente A"}).json()["id"]
    b = client.post("/api/clientes", json={"nombre": "Cliente B"}).json()["id"]

    r = client.post("/api/sectores", json={"cliente_id": a, "nombre": "Admision"})
    assert r.status_code == 201
    sector_id = r.json()["id"]
    client.post("/api/sectores", json={"cliente_id": a, "nombre": "Deposito"})
    client.post("/api/sectores", json={"cliente_id": b, "nombre": "Admision"})

    # El listado filtra por cliente: un sector es de un cliente y de uno solo.
    de_a = client.get(f"/api/sectores?cliente_id={a}").json()
    assert [s["nombre"] for s in de_a] == ["Admision", "Deposito"]
    assert len(client.get(f"/api/sectores?cliente_id={b}").json()) == 1
    assert len(client.get("/api/sectores").json()) == 3

    r = client.put(f"/api/sectores/{sector_id}", json={"nombre": "Admision y Guardia"})
    assert r.status_code == 200
    assert r.json()["nombre"] == "Admision y Guardia"

    assert client.delete(f"/api/sectores/{sector_id}").status_code == 204
    assert [s["nombre"] for s in client.get(f"/api/sectores?cliente_id={a}").json()] == ["Deposito"]


def test_sector_repetido_en_el_mismo_cliente_da_409_pero_no_entre_clientes(client):
    """El nombre es unico POR cliente: dos clientes distintos pueden tener
    los dos su "Administracion", que es el caso normal."""
    _login(client)
    a = client.post("/api/clientes", json={"nombre": "Cliente A"}).json()["id"]
    b = client.post("/api/clientes", json={"nombre": "Cliente B"}).json()["id"]

    assert client.post("/api/sectores", json={"cliente_id": a, "nombre": "Administracion"}).status_code == 201
    assert client.post("/api/sectores", json={"cliente_id": a, "nombre": "Administracion"}).status_code == 409
    assert client.post("/api/sectores", json={"cliente_id": b, "nombre": "Administracion"}).status_code == 201


def test_borrar_un_sector_desasigna_sus_incidencias_y_no_las_borra(client):
    """El `ondelete="SET NULL"` de `incidencias.sector_id` NO se ejecuta: el
    engine no activa `PRAGMA foreign_keys`. Sin la desasignacion explicita del
    repositorio, el ticket quedaba apuntando a un sector inexistente y el
    reporte de Incidencias --que resuelve el sector por join-- lo mostraba
    vacio sin forma de saber por que."""
    _login(client)
    cliente_id = client.post("/api/clientes", json={"nombre": "Cliente Test"}).json()["id"]
    sector_id = client.post(
        "/api/sectores", json={"cliente_id": cliente_id, "nombre": "Admision"}
    ).json()["id"]
    otro_sector = client.post(
        "/api/sectores", json={"cliente_id": cliente_id, "nombre": "Deposito"}
    ).json()["id"]

    con_sector = client.post("/api/incidencias", json={
        "cliente_id": cliente_id, "sector_id": sector_id, "titulo": "No imprime",
    }).json()["id"]
    intacta = client.post("/api/incidencias", json={
        "cliente_id": cliente_id, "sector_id": otro_sector, "titulo": "Sin red",
    }).json()["id"]

    assert client.delete(f"/api/sectores/{sector_id}").status_code == 204

    # La incidencia sigue existiendo -- borrar un sector no borra tickets.
    r = client.get(f"/api/incidencias/{con_sector}")
    assert r.status_code == 200
    assert r.json()["sector_id"] is None

    # Y no se lleva puesto el sector de las demas.
    assert client.get(f"/api/incidencias/{intacta}").json()["sector_id"] == otro_sector


def test_sector_inexistente_da_404(client):
    _login(client)
    assert client.put("/api/sectores/9999", json={"nombre": "X"}).status_code == 404
    assert client.delete("/api/sectores/9999").status_code == 404


def test_sectores_exige_sesion(client):
    """Todos los routers de dominio van detras de `require_staff` -- este
    tambien, que es la clase de cosa que ya se habia escapado una vez."""
    assert client.get("/api/sectores").status_code == 401
    assert client.post("/api/sectores", json={"cliente_id": 1, "nombre": "X"}).status_code == 401


# --- Borrados que tienen que limpiar lo suyo ------------------------------
#
# Los `ondelete` declarados en los modelos NO se ejecutan: el engine no activa
# `PRAGMA foreign_keys`. Cada repositorio lo hace explicito. Estos tests fijan
# ese contrato, que ademas es lo que prometen los dialogos de confirmacion de
# la UI -- la vez pasada el dialogo prometia una cascada que no ocurria.


def test_borrar_un_tecnico_desasigna_sus_incidencias_y_no_las_borra(client):
    _login(client)
    cliente_id = client.post("/api/clientes", json={"nombre": "Cliente Test"}).json()["id"]
    tecnico_id = client.post("/api/tecnicos", json={"nombre": "Mariano"}).json()["id"]
    otro = client.post("/api/tecnicos", json={"nombre": "Tomas"}).json()["id"]

    suya = client.post("/api/incidencias", json={
        "cliente_id": cliente_id, "tecnico_id": tecnico_id, "titulo": "No imprime",
    }).json()["id"]
    ajena = client.post("/api/incidencias", json={
        "cliente_id": cliente_id, "tecnico_id": otro, "titulo": "Sin red",
    }).json()["id"]

    assert client.delete(f"/api/tecnicos/{tecnico_id}").status_code == 204

    r = client.get(f"/api/incidencias/{suya}")
    assert r.status_code == 200
    assert r.json()["tecnico_id"] is None
    # No se lleva puesto el tecnico de las demas.
    assert client.get(f"/api/incidencias/{ajena}").json()["tecnico_id"] == otro


def test_borrar_un_equipo_borra_su_historial_y_desasigna_sus_incidencias(client):
    """El caso que ya se habia demostrado solo en dev el 2026-07-30: un script
    borro sus equipos de prueba por la API y los 10 movimientos sobrevivieron
    huerfanos."""
    _login(client)
    cliente_id = client.post("/api/clientes", json={"nombre": "Cliente Test"}).json()["id"]
    equipo_id = client.post("/api/equipos", json={
        "cliente_id": cliente_id, "tipo": "Impresora", "sector": "Admision",
    }).json()["id"]
    otro_equipo = client.post("/api/equipos", json={
        "cliente_id": cliente_id, "tipo": "Notebook",
    }).json()["id"]

    # El alta genera un movimiento, y el traslado otro.
    client.put(f"/api/equipos/{equipo_id}", json={
        "cliente_id": cliente_id, "tipo": "Impresora", "sector": "Deposito",
        "estado": "activo", "motivo": "Se movio",
    })
    assert len(client.get(f"/api/equipos/{equipo_id}/movimientos").json()) >= 2

    incidencia_id = client.post("/api/incidencias", json={
        "cliente_id": cliente_id, "equipo_id": equipo_id, "titulo": "No imprime",
    }).json()["id"]
    ajena = client.post("/api/incidencias", json={
        "cliente_id": cliente_id, "equipo_id": otro_equipo, "titulo": "Sin red",
    }).json()["id"]

    assert client.delete(f"/api/equipos/{equipo_id}").status_code == 204

    # El ticket sobrevive, sin equipo.
    r = client.get(f"/api/incidencias/{incidencia_id}")
    assert r.status_code == 200
    assert r.json()["equipo_id"] is None
    assert client.get(f"/api/incidencias/{ajena}").json()["equipo_id"] == otro_equipo

    # Y el historial del equipo borrado no sobrevive huerfano. El endpoint
    # filtra por equipo_id sin chequear que el equipo exista, asi que si
    # quedaran filas colgadas las devolveria igual -- es la comprobacion
    # directa de lo que paso en dev.
    assert client.get(f"/api/equipos/{equipo_id}/movimientos").json() == []
    # El historial del otro equipo queda intacto: se borro lo suyo y nada mas.
    assert len(client.get(f"/api/equipos/{otro_equipo}/movimientos").json()) >= 1


# --- Baja logica de clientes ----------------------------------------------
#
# Decidido el 2026-08-01: un cliente no se borra, se desactiva (como en
# Contalibra). Eso disuelve el problema de los huerfanos en vez de resolverlo.


def test_desactivar_y_reactivar_un_cliente_no_toca_su_historial(client):
    _login(client)
    cliente_id = client.post("/api/clientes", json={"nombre": "Cliente Test"}).json()["id"]
    equipo_id = client.post("/api/equipos", json={
        "cliente_id": cliente_id, "tipo": "Impresora",
    }).json()["id"]
    incidencia_id = client.post("/api/incidencias", json={
        "cliente_id": cliente_id, "titulo": "No imprime",
    }).json()["id"]

    r = client.post(f"/api/clientes/{cliente_id}/desactivar")
    assert r.status_code == 200
    assert r.json()["activo"] is False

    # Lo que importa: el historial sigue entero y apuntando al cliente.
    assert client.get(f"/api/equipos/{equipo_id}").json()["cliente_id"] == cliente_id
    assert client.get(f"/api/incidencias/{incidencia_id}").json()["cliente_id"] == cliente_id

    # Y el listado lo filtra o no segun se pida.
    assert cliente_id not in [c["id"] for c in client.get("/api/clientes?solo_activos=true").json()]
    assert cliente_id in [c["id"] for c in client.get("/api/clientes").json()]

    # Es reversible, que es la razon de ser de la baja logica.
    r = client.post(f"/api/clientes/{cliente_id}/activar")
    assert r.status_code == 200
    assert r.json()["activo"] is True
    assert cliente_id in [c["id"] for c in client.get("/api/clientes?solo_activos=true").json()]


def test_no_se_puede_borrar_un_cliente_con_historial(client):
    """El router SIEMPRE declaro este 409 en un `except IntegrityError`, pero
    esa rama no se ejecutaba nunca: sin `PRAGMA foreign_keys` la base no
    levanta el error y el DELETE pasaba, dejando todo huerfano. Ahora el
    chequeo es explicito."""
    _login(client)
    cliente_id = client.post("/api/clientes", json={"nombre": "Cliente Test"}).json()["id"]
    equipo_id = client.post("/api/equipos", json={
        "cliente_id": cliente_id, "tipo": "Impresora",
    }).json()["id"]

    r = client.delete(f"/api/clientes/{cliente_id}")
    assert r.status_code == 409
    assert "1 equipos" in r.json()["detail"]
    assert "esactiva" in r.json()["detail"]  # sugiere la salida correcta

    # El cliente y su equipo siguen ahi: no se borro nada a medias.
    assert client.get(f"/api/clientes/{cliente_id}").status_code == 200
    assert client.get(f"/api/equipos/{equipo_id}").status_code == 200


def test_un_cliente_vacio_si_se_puede_borrar(client):
    """El caso de uso que justifica conservar el DELETE: uno cargado por error."""
    _login(client)
    cliente_id = client.post("/api/clientes", json={"nombre": "Cargado por error"}).json()["id"]

    assert client.delete(f"/api/clientes/{cliente_id}").status_code == 204
    assert client.get(f"/api/clientes/{cliente_id}").status_code == 404


def test_activar_desactivar_un_cliente_inexistente_da_404(client):
    _login(client)
    assert client.post("/api/clientes/9999/desactivar").status_code == 404
    assert client.post("/api/clientes/9999/activar").status_code == 404
