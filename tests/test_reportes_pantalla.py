"""Los reportes en pantalla, y que sean **el mismo reporte** que el Excel
(2026-08-04).

Hasta ahora los seis analiticos solo se podian bajar; ahora `GET /<slug>`
devuelve la vista en JSON y `GET /<slug>.xlsx` la baja. El riesgo que introduce
esa segunda salida es la divergencia: que alguien agregue una columna de un
lado y no del otro, y queden dos reportes con el mismo nombre y distinto
contenido. Por eso el test central de este archivo **compara las dos salidas
celda por celda** en vez de mirarlas por separado.

Los otros dos que importan:

- Que los numeros sigan siendo numeros en el Excel. La vista arma texto para la
  pantalla; si ese texto se escribiera tal cual en la planilla, la columna "#"
  y los conteos dejarian de sumarse y de ordenarse, que es para lo que se baja
  el archivo.
- Que los resaltados viajen como marca semantica y no como color: es lo que
  permite que el Excel los pinte y la pantalla les ponga una clase.
"""
from datetime import date, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook

# `client` sale de conftest.py.


def _iso(dias: int) -> str:
    return (date.today() + timedelta(days=dias)).isoformat()


@pytest.fixture
def escenario(client):
    assert client.post(
        "/auth/login", json={"username": "admin", "password": "admin"}
    ).status_code == 200

    cliente_id = client.post("/api/clientes", json={
        "nombre": "Compulibra", "empresa": "Compulibra SRL", "email": "c@test.com",
    }).json()["id"]
    taller = client.post("/api/depositos", json={"nombre": "Taller"}).json()["id"]

    # Uno con la garantía vencida (se resalta), uno guardado en el depósito y
    # uno normal: los tres casos que el reporte de Equipamiento distingue.
    vencido = client.post("/api/equipos", json={
        "cliente_id": cliente_id, "tipo": "Notebook", "marca": "Lenovo",
        "serial": "S-1", "sector": "Admisión", "garantia_vence": _iso(-30),
    }).json()["id"]
    client.post("/api/equipos", json={
        "cliente_id": cliente_id, "tipo": "Impresora", "marca": "HP",
        "serial": "S-2", "sector": "Ventas",
    })
    guardado = client.post("/api/equipos", json={
        "cliente_id": cliente_id, "tipo": "Monitor", "marca": "Dell",
        "serial": "S-3", "sector": "Depósito viejo", "estado": "almacenado",
    }).json()["id"]
    client.post("/api/depositos/transferir", json={
        "equipo_ids": [guardado], "destino_id": taller,
    })

    client.post("/api/incidencias", json={
        "cliente_id": cliente_id, "equipo_id": vencido,
        "titulo": "No arranca", "prioridad": "alta",
    })

    return {"cliente_id": cliente_id, "taller": taller, "vencido": vencido}


def _celdas_del_xlsx(contenido: bytes) -> list[list]:
    """Las filas del Excel desde la fila de headers (la 4) en adelante."""
    ws = load_workbook(BytesIO(contenido)).active
    return [
        [c.value for c in fila]
        for fila in ws.iter_rows(min_row=4)
    ]


def _celdas_de_la_vista(vista: dict) -> list[list]:
    """Lo mismo pero desde el JSON, en el orden en que lo dibuja la pantalla:
    headers, después cada grupo (con su etiqueta si la tiene) y los totales."""
    filas = [[c["label"] for c in vista["columnas"]]]
    for grupo in vista["grupos"]:
        if grupo["etiqueta"]:
            filas.append([grupo["etiqueta"]])
        filas.extend(
            [c["texto"] if c["texto"] is not None else "—" for c in fila]
            for fila in grupo["filas"]
        )
    if vista["totales"]:
        filas.append([c["texto"] for c in vista["totales"]])
    return filas


# --- la pantalla y el Excel son el mismo reporte ----------------------------

@pytest.mark.parametrize("ruta", [
    "equipamiento",
    "garantias?dias=60",
    "movimientos?desde=2020-01-01&hasta=2099-12-31",
    "incidencias-periodo?desde=2020-01-01&hasta=2099-12-31",
    "tecnico?desde=2020-01-01&hasta=2099-12-31",
    "facturacion?desde=2020-01-01&hasta=2099-12-31",
    "clientes",
    "equipos",
    "incidencias",
])
def test_la_pantalla_y_el_excel_traen_lo_mismo(client, escenario, ruta):
    """El test que justifica haber extraido la vista: si alguien agrega una
    columna a una sola de las dos salidas, esto se pone en rojo."""
    slug, _, query = ruta.partition("?")
    sufijo = f"?{query}" if query else ""

    vista = client.get(f"/api/reportes/{slug}{sufijo}").json()
    excel = client.get(f"/api/reportes/{slug}.xlsx{sufijo}")

    assert excel.status_code == 200
    del_excel = _celdas_del_xlsx(excel.content)
    de_la_vista = _celdas_de_la_vista(vista)

    assert len(del_excel) == len(de_la_vista)
    for fila_excel, fila_vista in zip(del_excel, de_la_vista):
        # El Excel escribe la fila completa (con celdas vacías a la derecha en
        # los encabezados de grupo); la vista sólo lo que tiene texto.
        recortada = fila_excel[:len(fila_vista)]
        assert [
            str(v) if v is not None else None for v in recortada
        ] == [
            str(v) if v is not None else None for v in fila_vista
        ]


# --- lo que cada salida hace distinto, a propósito ---------------------------

def test_los_numeros_siguen_siendo_numeros_en_el_excel(client, escenario):
    """La vista arma texto para la pantalla. Si ese texto llegara tal cual a la
    planilla, la columna "#" no se sumaria ni se ordenaria."""
    excel = client.get(
        "/api/reportes/incidencias-periodo.xlsx?desde=2020-01-01&hasta=2099-12-31"
    )

    ws = load_workbook(BytesIO(excel.content)).active
    primera_fila_de_datos = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]

    assert isinstance(primera_fila_de_datos[0], int)


def test_los_resaltados_viajan_como_marca_y_no_como_color(client, escenario):
    """La pantalla no puede recibir `FFFEE2E2`: cada salida traduce la marca a
    lo suyo (un relleno en el Excel, una clase en la tabla)."""
    vista = client.get("/api/reportes/equipamiento").json()

    filas = vista["grupos"][0]["filas"]
    marcas = {c["marca"] for fila in filas for c in fila if c["marca"]}

    assert "peligro" in marcas  # la garantía vencida
    assert all(not str(m).startswith("FF") for m in marcas)


def test_el_equipo_guardado_muestra_el_deposito_y_no_el_sector_viejo(client, escenario):
    """La columna dice "Sector / Depósito" justamente por esto: un equipo en el
    taller no está en ningún sector del cliente."""
    vista = client.get("/api/reportes/equipamiento").json()

    columnas = [c["label"] for c in vista["columnas"]]
    lugar = columnas.index("Sector / Depósito")
    lugares = {fila[lugar]["texto"] for fila in vista["grupos"][0]["filas"]}

    assert "Taller" in lugares
    assert "Depósito viejo" not in lugares


def test_la_vista_declara_los_filtros_aplicados(client, escenario):
    """Es lo que encabeza el Excel y lo que imprime la pantalla: una hoja
    suelta sin decir de qué cliente y de qué período es no sirve."""
    vista = client.get(
        f"/api/reportes/equipamiento?cliente_id={escenario['cliente_id']}&estado=activo"
    ).json()

    assert "Cliente: Compulibra SRL" in vista["filtros"]
    assert "Estado: Activo" in vista["filtros"]


def test_facturacion_viene_agrupada_por_cliente(client, escenario):
    """El único de los seis con grupos. Si la agrupación se perdiera, la vista
    seguiría teniendo las filas y nadie lo notaría hasta mirar el Excel."""
    # Se abre y después se cierra: el reporte filtra por `fecha_cierre`, que la
    # pone el update al pasar a cerrado. Naciendo cerrada no tendría fecha y
    # quedaría fuera de todo período.
    incidencia_id = client.post("/api/incidencias", json={
        "cliente_id": escenario["cliente_id"], "titulo": "Cerrada",
    }).json()["id"]
    client.put(f"/api/incidencias/{incidencia_id}", json={
        "cliente_id": escenario["cliente_id"], "titulo": "Cerrada", "estado": "cerrado",
    })

    vista = client.get(
        "/api/reportes/facturacion?desde=2020-01-01&hasta=2099-12-31"
    ).json()

    assert len(vista["grupos"]) == 1
    assert vista["grupos"][0]["etiqueta"].startswith("Compulibra SRL — 1 incidencia")


def test_un_reporte_sin_datos_no_rompe(client):
    """Base recién creada: la vista tiene columnas y cero filas, y el Excel
    igual se genera."""
    client.post("/auth/login", json={"username": "admin", "password": "admin"})

    vista = client.get("/api/reportes/equipamiento").json()

    assert vista["cantidad_filas"] == 0
    assert len(vista["columnas"]) > 0
    assert client.get("/api/reportes/equipamiento.xlsx").status_code == 200
