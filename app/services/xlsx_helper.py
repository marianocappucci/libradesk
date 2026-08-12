"""Helpers de export a Excel con openpyxl — reconstruye el diseno visual
de `xlsxHelper.ts` (backend Node.js viejo) en Python: paleta indigo,
encabezado con titulo+filtros, header congelado, filas alternadas,
resaltado por celda, fila de totales y encabezados de grupo. No hay
precedente de export xlsx en la familia Libra (Gestiolibra no exporta) —
es una pieza propia de LibraDesk.

**Solo el como, no el que.** Las etiquetas de estado/prioridad/cobro y los
colores de resaltado vivian aca hasta el 2026-08-04; se mudaron a
`reporte_vista.py` cuando los reportes empezaron a verse tambien en pantalla,
porque dejaron de ser del Excel. Este modulo quedo con lo que si es suyo:
poner celdas en una hoja. Quien lo llama es `reporte_xlsx.py`."""
from datetime import datetime
from io import BytesIO

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

PRIMARY = "FF4338CA"  # indigo-700
HDR_BG = "FFE0E7FF"  # indigo-100
ALT_ROW = "FFFAFAFA"
WHITE = "FFFFFFFF"
GROUP_BG = "FFEEF2FF"
GROUP_FG = "FF3730A3"
HDR_FG = "FF1E1B4B"
BORDE_HDR = "FFC7D2FE"

_HAIR = Border(bottom=Side(style="hair", color="FFEDEDED"))
_BOTTOM_HDR = Border(bottom=Side(style="thin", color=BORDE_HDR))
_TOP_HDR = Border(top=Side(style="thin", color=BORDE_HDR))


def _fill(argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb)


def create_sheet(titulo: str, filtros: list[str], sheet_name: str | None = None):
    """Devuelve (wb, ws) con el encabezado de titulo+filtros ya puesto.
    La primera fila util (la de headers) es la 4."""
    wb = Workbook()
    wb.properties.creator = "LibraDesk"
    ws = wb.active
    ws.title = (sheet_name or titulo)[:31]
    # Apaisado y ajustado al ancho: estos reportes tienen muchas columnas.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return wb, ws


def add_meta_header(ws: Worksheet, titulo: str, filtros: list[str], col_count: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    c1 = ws.cell(row=1, column=1, value=f"LibraDesk — {titulo}")
    c1.font = Font(bold=True, size=13, color=WHITE)
    c1.fill = _fill(PRIMARY)
    c1.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 26

    info = "   |   ".join(
        [f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}", *filtros]
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    c2 = ws.cell(row=2, column=1, value=info)
    c2.font = Font(size=9, italic=True, color="FF6B7280")
    c2.fill = _fill("FFF1F5F9")
    c2.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 5
    return 4


def add_header_row(ws: Worksheet, row_num: int, headers: list[str], widths: list[int]) -> None:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=i, value=h)
        cell.font = Font(bold=True, size=10, color=HDR_FG)
        cell.fill = _fill(HDR_BG)
        cell.border = _BOTTOM_HDR
        cell.alignment = Alignment(vertical="center", horizontal="left")
        ws.column_dimensions[get_column_letter(i)].width = (
            widths[i - 1] if i - 1 < len(widths) else 14
        )
    ws.row_dimensions[row_num].height = 20
    ws.freeze_panes = f"A{row_num + 1}"


def add_data_row(
    ws: Worksheet,
    row_num: int,
    values: list,
    cell_fills: list[str | None] | None = None,
    is_alt: bool = False,
) -> None:
    row_bg = ALT_ROW if is_alt else WHITE
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=i, value=v if v is not None else "—")
        cell.font = Font(size=10)
        propio = cell_fills[i - 1] if cell_fills and i - 1 < len(cell_fills) else None
        cell.fill = _fill(propio or row_bg)
        cell.alignment = Alignment(vertical="center")
        cell.border = _HAIR
    ws.row_dimensions[row_num].height = 16


def add_totals_row(ws: Worksheet, row_num: int, values: list) -> None:
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=i, value=v if v is not None else "")
        cell.font = Font(bold=True, size=10, color=HDR_FG)
        cell.fill = _fill(HDR_BG)
        cell.border = _TOP_HDR
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row_num].height = 18


def add_group_header(ws: Worksheet, row_num: int, text: str, col_count: int) -> None:
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font = Font(bold=True, size=10, color=GROUP_FG)
    cell.fill = _fill(GROUP_BG)
    cell.border = _TOP_HDR
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row_num].height = 18


def xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
